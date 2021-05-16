import tkinter
from tkinter import *
from tkinter import ttk
from tkinter.ttk import *
from tkinter.filedialog import askopenfilename
import os
import time
from datetime import datetime, timedelta
import csv
from PIL import ImageTk, Image

from openpyxl import load_workbook
import threading
import re


# Define the attribute set code specified in the sheet or use deafult one
def attribute_set_code(row, row_default):
    wb = load_workbook(attributes_filename)
    sheet = wb["attribute_sets"]
    for attr_row in sheet.rows:
        if attr_row[0].value.upper() in row[4].upper():
            row[10], row[11], row[12], row[13], row[14], row[15] = attr_row[0].value, attr_row[
                1].value, attr_row[2].value, attr_row[3].value, attr_row[4].value, attr_row[5].value
            break
        else:
            row[10], row[11], row[12], row[13], row[14], row[15] = sheet['A'+row_default].value, sheet['B'+row_default].value, sheet['C' +
                                                                                                                                     row_default].value, sheet['D'+row_default].value, sheet['E'+row_default].value, sheet['F'+row_default].value


# This function adds all other properties from Properties file
def add_properties(row, col_num, col_name, choose_category):

    # Read All Excell Sheets
    wb = load_workbook(attributes_filename)

    brand = wb["brand"]
    processor_type = wb["processor_type"]
    memorysize = wb["memorysize"]
    disk_size = wb["disk_size"]
    operating_system = wb["operating_system"]
    security_options = wb["security_options"]
    display_resolutions = wb["display_resolutions"]
    display_type = wb["display_type"]
    displaysize = wb["displaysize"]
    categories = wb["categories"]

    # Specify brand depends on the mapping
    for brand_row in brand:
        if brand_row[0].value.upper() == row[1].upper():
            row[1] = brand_row[0].value
            break

    # Apply attributes for Desktops and Laptops
    if row[col_num] == choose_category[0] or row[col_num] == choose_category[1]:
        for proc_row in processor_type:
            if proc_row[0].value.split(' ')[1].upper() in row[col_name].upper():
                row[24] = proc_row[0].value
                break
        for mem_row in memorysize:
            if re.search(fr'\D{mem_row[0].value}', row[col_name], re.I):
                row[25] = mem_row[0].value
                break
        for disk_row in disk_size:
            if disk_row[0].value in row[col_name]:
                row[26] = disk_row[0].value
                break
        for sys_row in operating_system:
            os = sys_row[1].value.split(",")
            for i in os:
                if i in row[col_name].upper():
                    row[27] = sys_row[0].value
                    break
                break
        for sec_row in security_options:
            if sec_row[0].value.upper() in row[col_name].upper():
                row[28] = sec_row[0].value
                break

    # Apply attributes for Monitors
    elif row[col_num] == choose_category[2]:
        for dis_row in display_resolutions:
            if dis_row[0].value.upper() in row[col_name].upper():
                row[29] = dis_row[0].value
                break
        for dis_row in display_type:
            if dis_row[0].value.upper() in row[col_name].upper():
                row[30] = dis_row[0].value
                break
        for dis_row in displaysize:
            if dis_row[0].value.upper() in row[col_name].upper():
                row[31] = dis_row[0].value
                break
    else:
        pass

    # Apply Category (if category is not found default one is used) - defult category does not work yet
    for cat_row in categories:
        if row[col_num].upper() in cat_row[0].value.upper().split(","):
            row[col_num] = cat_row[1].value
            break
        # else:
        #     row[col_num] = categories['B6'].value
        #     break


# Main function that stores all the logic
def process_csv(suppliers, app):
    result_filename = datetime.now().strftime("%c")
    result_filename = result_filename.replace(":", ".", 2)

    # Getting supplier name from user's input
    supplier_name = suppliers.get()

    if supplier_name in ['Dicker Data Ltd - AU', 'Dicker Data Ltd - NZ', 'Dove Electronics', 'Ingram Micro NZ LTD', 'MacGear New Zealand']:

        # Open file and and specifing delimeter for files
        with open(main_filename, 'r', encoding='utf8', errors='replace') as csv_file:
            if supplier_name == 'Dove Electronics' or supplier_name == 'MacGear New Zealand':
                csv_reader = csv.reader(csv_file, delimiter=';')
            else:
                csv_reader = csv.reader(csv_file, delimiter=',')

            # Skip the header, because do not need to read it
            next(csv_reader)

            # Create csv result file
            with open(f'{supplier_name + " " + result_filename}.csv', 'w', newline='', encoding='utf8', errors='replace') as result_csv_file:
                csv_writer = csv.writer(result_csv_file, delimiter=',')

                # Declaring headers and writing to the result.csv file
                headers = ['sku', 'brand', 'vpn', 'name', 'categories', 'price', 'cost', 'special_price', 'qty', 'supplier', 'attribute_set_code', 'weight', 'shipment_type_nzpost', 'dimensions_height',
                           'dimensions_length', 'dimensions_width', 'product_website', 'tax_class_name', 'visibility', 'base_image', 'base_image_label', 'small_image', 'small_image_label', 'thumbnail_image', 'processor_type', 'memorysize', 'disk_size', 'operating_system', 'security_options', 'display_resolutions', 'display_type', 'displaysize', 'description', 'product_type']

                csv_writer.writerow(headers)

                # Do all processing for each row of the input file
                for row in csv_reader:

                    # array for data handling
                    array = []

                    # Define the data processing flow basing on specific supplier
                    if supplier_name == 'Dicker Data Ltd - AU' or supplier_name == 'Dicker Data Ltd - NZ':

                        # adding empty items for each row for assinging new values
                        row.extend((" " * 17).split(" "))

                        # remove redundant columns
                        del row[5:7]

                        row[5] = row[5].replace(",", "")
                        row[6] = row[6].replace(",", "")

                        # Convert price and cost values for DDAU supplier
                        if supplier_name == 'Dicker Data Ltd - AU':
                            row[5] = float(row[5]) * 1.12
                            row[6] = float(row[6]) * 1.12
                        else:
                            pass
                        row[8] = round((float(row[6]) / 0.9), 2)
                        row[7], row[8] = row[8], row[7]

                        filter_by_category = ["SOFTWARE", "WARRANTY", "STORAGE MEDIA",
                                              "CONSUMABLES", "CLEARANCE", "Unknown Primary Category", ""]
                        filter_by_qty = ["0", "999999999"]

                        if (row[4] not in filter_by_category) and (row[8] not in filter_by_qty):

                            # Add supplier name basing on user choice
                            row[9] = supplier_name

                            # Think about attribute_set_code dynamic architecture
                            attribute_set_code(row, '8')

                            row[16], row[17], row[18] = "base", "Taxable Goods", "Catalog, Search"

                            add_properties(
                                row, 4, 3, ['DESKTOPS', 'NOTEBOOKS', 'MONITORS'])

                            del row[-1]
                            row.append('simple')

                            csv_writer.writerow(row)
                        else:
                            pass
                    elif supplier_name == 'Dove Electronics':

                        # Filter categories
                        filter_by_category = ["_Unspecified", "Unspecified", "Adapters", "Batteries & Chargers", "Binoculars",
                                              "Cables – Other", "Cables - Monitor & Video", "Cables - USB & Firewire", "Calculators", "Cases - OEM PC", "Cleaning", "Clearance", "Clearance", "CPU", "Fax Machines", "Fans", "Furniture & Ergonomics", "IO Cards", "Motherboards", "Refurbished", "Server Cabinets", "Servers", "Software – Applications", "Sewing & Craft", "Speakers", "Security – DVR", "Security – Cameras", "Software - Operating Systems"]

                        if (row[8] not in filter_by_category) and (row[7] != "0"):

                            row[5] = float(
                                (row[5].replace(",", ".")[1:]).replace(" ", ""))
                            row[6] = float(
                                (row[6].replace(",", ".")[1:]).replace(" ", ""))
                            special_price = round((float(row[6]) / 0.9), 2)

                            base_image = row[11]

                            array.extend([row[4], row[0], row[1], row[2], row[8], row[5],
                                          row[6], special_price, int(row[7]), supplier_name])

                            array.extend((' ' * 5).split(' '))

                            attribute_set_code(array, '8')

                            array.extend(
                                ["base", "Taxable Goods", "Catalog, Search", base_image, "", "", "", ""])

                            array.extend((' ' * 7).split(' '))

                            add_properties(
                                array, 4, 3, ["Desktop PC's", "Notebooks / Laptops", "Monitors"])

                            # add description and product_type to the end
                            array.extend([row[3], 'simple'])

                            csv_writer.writerow(array)
                        else:
                            pass
                    elif supplier_name == 'Ingram Micro NZ LTD':

                        # Filter categories
                        filter_by_category = [
                            "Ingram Micro Logistics", "Jewelry", "Physical Security", "Software", "Training"]

                        if (row[87] not in filter_by_category) and (row[16] != "0"):

                            row[13] = float(row[13])
                            row[14] = float(row[14])
                            special_price = round((float(row[13]) / 0.9), 2)

                            base_image = ""

                            array.extend([row[0], row[7], row[6], row[1], row[88], row[14],
                                          row[13],  special_price, int(row[16]), supplier_name])

                            array.extend((' ' * 5).split(' '))

                            attribute_set_code(array, '8')

                            array.extend(
                                ["base", "Taxable Goods", "Catalog, Search", base_image, "", "", "", ""])

                            array.extend((' ' * 7).split(' '))

                            add_properties(
                                array, 4, 3, ["Desktops", "Notebooks & Tablets", "Monitors"])

                            # add description and product_type to the end
                            array.extend([row[23], 'simple'])

                            csv_writer.writerow(array)
                        else:
                            pass
                    elif supplier_name == 'MacGear New Zealand':

                        # Process only for quantity more than 0
                        if row[4] != "0":

                            price = float(row[5].replace(",", "."))
                            cost = float(row[6].replace(",", "."))
                            special_price = round((cost / 0.9), 2)

                            base_image = row[12]

                            array.extend(
                                [row[0], row[2], row[7], row[1], "", price, cost, special_price, row[4], supplier_name])

                            array.extend((' ' * 5).split(' '))

                            attribute_set_code(array, '8')
                            # array[10], array[11], array[12], array[13], array[14], array[15] = "", row[8], "", "", "", ""

                            array.extend(
                                ["base", "Taxable Goods", "Catalog, Search", base_image, "", "", "", ""])

                            array.extend((' ' * 7).split(' '))

                            add_properties(array, 4, 3, ["", "", ""])
                            # array[24], array[25], array[26], array[27], array[28], array[29], array[30], array[31] = "", "", "", "", "", "", "", ""

                            # add description and product_type to the end
                            array.extend(['', 'simple'])

                            csv_writer.writerow(array)
                        else:
                            pass
                    else:
                        break

                finish_ok(app)
    else:
        # Show notification that no setting for supplier
        finish(app)

    print("done")


# Compare files main function
def compare_files(suppliers, app):
    result_filename = datetime.now().strftime("%c")
    result_filename = result_filename.replace(":", ".", 2)

    supplier_name = suppliers.get()

    if supplier_name in ['Dicker Data Ltd - AU', 'Dicker Data Ltd - NZ', 'Dove Electronics', 'Ingram Micro NZ LTD', 'MacGear New Zealand']:
        # Get price & qty data from result file
        def price_qty_data(sku_col_res, price_col_res, qty_col_res, round_price):
            with open(result_file, 'r') as csv_result_file:
                csv_result_reader = csv.reader(csv_result_file, delimiter=',')
                next(csv_result_reader)

                price = {}
                qty = {}
                for row in csv_result_reader:
                    if round_price:
                        price.update(
                            {row[sku_col_res].strip().upper(): round_price(row[price_col_res])})
                    else:
                        price.update(
                            {row[sku_col_res].strip().upper(): float(row[price_col_res].strip())})
                    qty.update(
                        {row[sku_col_res].strip().upper(): int(row[qty_col_res].strip())})
            return price, qty

        # Rounding price result file
        def round_price(price):
            return round(float(price.strip()), 2)

        # Price formatting for DDAU
        def format_price_1(price):
            return round(float((price.replace(",", "")).strip()) * 1.12, 2)

        # Price formatting for DDNZ
        def format_price_2(price):
            return round(float((price.replace(",", "")).strip()), 2)

        # Price formatting for Dove supplier
        def format_price_3(price):
            return float((price.replace(",", ".")[1:]).replace(" ", ""))

        # Price formatting for All other suppliers
        def format_price_all(price):
            return float((price.replace(",", ".")).strip())

        # Compare price & qty and write to file
        def write_comparison(sku_col_res, price_col_res, qty_col_res, sku_col_raw, price_col_raw, qty_col_raw, price_format, round_price=None):
            price, qty = price_qty_data(
                sku_col_res, price_col_res, qty_col_res, round_price)
            for row in csv_reader:
                array = []

                row[price_col_raw] = price_format(row[price_col_raw])
                row[qty_col_raw] = int(row[qty_col_raw])
                if row[sku_col_raw].strip().upper() in price:
                    if row[price_col_raw] != price[row[sku_col_raw].strip()] or row[qty_col_raw] != qty[row[sku_col_raw].strip()]:
                        # Write data into the result file
                        array.extend(
                            [row[sku_col_raw], row[price_col_raw], row[qty_col_raw]])
                        csv_writer.writerow(array)
                    else:
                        pass
                        # print("No changed price or qty values")
                else:
                    pass
                    # print("No SKU found")
            print('Files compared')

        # Create result file
        with open(main_filename, 'r') as csv_file, open(f'{supplier_name + " " + result_filename}.csv', 'w', newline='') as result_csv_file:
            if supplier_name == 'Dove Electronics' or supplier_name == 'MacGear New Zealand':
                csv_reader = csv.reader(csv_file, delimiter=';')
            else:
                csv_reader = csv.reader(csv_file, delimiter=',')

            next(csv_reader)
            csv_writer = csv.writer(result_csv_file, delimiter=',')

            headers = ['sku', 'price', 'qty']
            csv_writer.writerow(headers)

            if supplier_name == 'Dicker Data Ltd - AU':
                write_comparison(0, 5, 8, 0, 7, 9, format_price_1, round_price)
            elif supplier_name == 'Dicker Data Ltd - NZ':
                write_comparison(0, 5, 8, 0, 7, 9, format_price_2, round_price)
            elif supplier_name == 'Dove Electronics':
                write_comparison(0, 5, 8, 4, 5, 7, format_price_3)
            elif supplier_name == 'Ingram Micro NZ LTD':
                write_comparison(0, 5, 8, 0, 14, 16, format_price_all)
            elif supplier_name == 'MacGear New Zealand':
                write_comparison(0, 5, 8, 0, 5, 4, format_price_all)
            else:
                print('Please check supplier name!')
        finish_ok(app)
    else:
        finish(app)


# Show notification that processing has ended OK
def finish_ok(app):
    finish_ok_label = ttk.Label(
        app, text="The file has been processed!", style='FinishOk.TLabel')
    finish_ok_label.place(x=153, y=350)
    print("The file has been processed!")


# Show "No settings for supplier" notification
def finish(app):
    finish_label = ttk.Label(
        app, text="There is no settings for this supplier! \n Please check supplier name!", style='Finish.TLabel')
    finish_label.place(x=150, y=350)
    print('There is no settings for this supplier!')


# GUI function that runs the app and support all logic
def run_app():
    app = Tk()
    app.title('Cool app')
    app.geometry("500x500")
    app.resizable(0, 0)

    # Style configs
    style = Style()
    style.configure('Start.TButton', font=(
        'calibri', 10, 'bold'), foreground='green')
    style.configure('Stop.TButton', font=(
        'calibri', 10, 'bold'), foreground='red')
    style.configure('FinishOk.TLabel', font=(
        'Times New Roman', 12, 'bold'), foreground='green')
    style.configure('Finish.TLabel', font=(
        'Times New Roman', 12, 'bold'), foreground='red')
    style.configure('Headers.TLabel', font=("Times New Roman", 11, 'bold'))

    # Open main file

    def main_file():
        global main_filename
        main_filename = askopenfilename(
            initialdir="./", title="Select CSV file", filetypes=[('Only CSV Files', '*.csv')])
        if not main_filename:
            pass
        else:
            short_name = os.path.basename(main_filename)
            label = ttk.Label(label_frame1, text=short_name)
            label.grid(column=1, row=2)
            print(main_filename)

    # Open result file

    def result_file():
        global result_file
        result_file = askopenfilename(
            initialdir="./", title="Select CSV file", filetypes=[('Only CSV Files', '*.csv')])
        if not result_file:
            pass
        else:
            short_name = os.path.basename(result_file)
            label = ttk.Label(label_frame3, text=short_name)
            label.grid(column=1, row=2)
            print(result_file)

    # Open attributes file

    def attributes_file():
        global attributes_filename
        attributes_filename = askopenfilename(
            initialdir="./", title="Select CSV file", filetypes=[('Only xlsx Files', '*.xlsx')])
        if not attributes_filename:
            pass
        else:
            # Open Suppliers sheet, get data from it
            wb = load_workbook(attributes_filename)
            sheet = wb["suppliers"]
            values = []
            # cells_coordinates = []
            # suppliers = {}
            for row in sheet.rows:
                for cell in row:
                    values.append(cell.value)
                    # cells_coordinates.append(cell.coordinate)
                    # suppliers[f"{cell.coordinate}"] = cell.value
            suppliers_values(values)
            # print(suppliers)

            short_name = os.path.basename(attributes_filename)
            label = ttk.Label(label_frame2, text="")
            label.grid(column=1, row=2)
            label.configure(text=short_name)
            print(attributes_filename)

    # Placing Logo image
    image = Image.open('logo.png')
    resize_image = image.resize((100, 30))
    img = ImageTk.PhotoImage(resize_image)
    my_img = Label(image=img)
    my_img.image = img
    my_img.place(x=280, y=300)

    # Label and Button for the main file
    label_frame1 = ttk.LabelFrame(app, text='Choose an input file', width=2)
    label_frame1.place(x=40, y=50)
    # grid(column=0, row=1, padx=20, pady=20)
    btn_file1 = ttk.Button(
        label_frame1, text="Browse a file", command=main_file)
    btn_file1.grid(column=1, row=1)

    # Label and Button for the attributes file
    label_frame2 = ttk.LabelFrame(app, text='Choose an Attributes file')
    label_frame2.place(x=40, y=132)
    # grid(column=0, row=2, padx=20, pady=20)
    btn_file2 = ttk.Button(
        label_frame2, text="Browse a file", command=attributes_file)
    btn_file2.grid(column=1, row=1)

    # Label and Button for the result file
    label_frame3 = ttk.LabelFrame(app, text='Choose the result file')
    label_frame3.place(x=300, y=50)
    btn_file3 = ttk.Button(
        label_frame3, text="Browse a file", command=result_file)
    btn_file3.grid(column=1, row=1)

    # Dropdown for attributes file
    ttk.Label(app, text="Select the supplier:", font=(
        "Times New Roman", 10)).place(x=90, y=220)
    n = tkinter.StringVar()

    suppliers = ttk.Combobox(app, width=27, textvariable=n, state="readonly")
    suppliers.place(x=200, y=220)

    # Adding values to drop down list
    def suppliers_values(values):
        suppliers['values'] = values

    # Thread for processing raw file
    def do_threading():
        main_thread = threading.Thread(
            target=process_csv, args=[suppliers, app])
        main_thread.start()

    # Thread for comapring raw and result files
    def do_threading1():
        main_thread = threading.Thread(
            target=compare_files, args=[suppliers, app])
        main_thread.start()

    # Choose what threads to use basing on the radiobutton choice
    def choose_process():
        if (i.get() == 1):
            do_threading()
        else:
            do_threading1()

    # Start & Stop Buttons (Stop button currently not used, because there is no optimal way to stop threads)
    btn_start = Button(app, text='Start', style='Start.TButton',
                       width=10, command=choose_process)
    # btn_stop = Button(app, text='Stop', style='Stop.TButton', width=10, command=app.destroy)

    # Place buttons
    btn_start.place(x=130, y=300)
    # btn_stop.place(x=290, y=300)

    # Radiobuttons buttons
    i = IntVar()
    radio_btn1 = Radiobutton(app, text="Process raw file", value=1, variable=i)
    radio_btn2 = Radiobutton(
        app, text="Compare Price & Quantity", value=2, variable=i)

    radio_btn1.place(x=290, y=135)
    radio_btn2.place(x=290, y=155)

    # GUI works by running infinite loop
    mainloop()


# Run app
run_app()
